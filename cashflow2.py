import xlrd
from xlrd.sheet import Hyperlink, Sheet
import time
import datetime
import numpy as np
import openpyxl
from classes import Secutiry, Timestamp, FloatRateEx, RiskfreeRate

def TakeThird(l: list):
    return l[2]

# 计算两天相差的年数
def CalcDateGap(date1: str, date2: str):
    date1=time.strptime(date1,"%Y/%m/%d")
    date2=time.strptime(date2,"%Y/%m/%d")
    date1=datetime.datetime(date1[0],date1[1],date1[2])
    date2=datetime.datetime(date2[0],date2[1],date2[2])
    #返回两个变量相差的值，就是相差天数
    return (date2 - date1).days / 365

# 从excel表格中得到浮动利率所需的信息
def GetFloatRateExtraData(excelData: Sheet) -> FloatRateEx:
    floatRate = FloatRateEx()
    # [1:]：去掉第一行汉字
    a = excelData.col_values(17)[1:]
    oneYearTime = []
    for it in a:
        if it != '':
            oneYearTime.append(xlrd.xldate_as_datetime(it, 0).strftime('%Y/%m/%d'))
    oneYearRate = excelData.col_values(18)[1:]

    a = excelData.col_values(20)[1:]
    SHIBOR1YearTime = []
    for it in a:
        if it != '':
            SHIBOR1YearTime.append(xlrd.xldate_as_datetime(it, 0).strftime('%Y/%m/%d'))
    SHIBOR1YearRate = excelData.col_values(21)[1:]

    a = excelData.col_values(23)[1:]
    SHIBOR3MonthTime = []
    for it in a:
        if it != '':
            SHIBOR3MonthTime.append(xlrd.xldate_as_datetime(it, 0).strftime('%Y/%m/%d'))
    SHIBOR3MonthRate = excelData.col_values(24)[1:]

    for i in range(len(oneYearTime)):
        floatRate.Add('oneYear', Timestamp(oneYearTime[i]), oneYearRate[i])
    for i in range(len(SHIBOR1YearTime)):
        floatRate.Add('SHIBOR1Year', Timestamp(SHIBOR1YearTime[i]), SHIBOR1YearRate[i])
    for i in range(len(SHIBOR3MonthTime)):
        floatRate.Add('SHIBOR3Month', Timestamp(SHIBOR3MonthTime[i]), SHIBOR3MonthRate[i])
    return floatRate

# 获得特定估值日期中债国债收益率
def GetBaseEarningRate(excelData: Sheet, date):
    riskfreeRate = RiskfreeRate()
    matus = excelData.col_values(1)[2:]
    rates = excelData.col_values(2)[2:]
    for i in range(len(matus)):
        riskfreeRate.Add(matus[i], rates[i] / 100)
    return riskfreeRate

# 使用二分法计算z-spread中的z值
# closingPrice: 收盘全价
# date: 估值日期
# riskfreeRate: 国债收益率数据
def CalcZSpread(closingPrice: float, date: str, timestamps, cashflow, riskfreeRate: RiskfreeRate) -> float:
    lb, rb = -0.2, 0.2
    # gap存放估值日期到每期现金流发生日的时间，单位：年
    # 形式为：[0.8, 1.8, 2.8, 3.8]
    gap = []
    for i in range(len(timestamps)):
        gap.append(CalcDateGap(date, timestamps[i].__str__()))
    while rb - lb > 0.00000001:
        mid = (rb + lb) / 2
        value = sum([cashflow[i] / ((1 + mid + riskfreeRate.GetRate(gap[i])) ** gap[i])
            for i in range(len(cashflow))])
        if value > closingPrice:
            lb = mid
        else:
            rb = mid
    return lb

# 返回excel表格数据，债券信息，浮动利率数据，国债数据
def ReadExcelData(excelFilePath):
    data = xlrd.open_workbook(excelFilePath)
    baseData = data.sheet_by_name(u'1_特定估值日期数据')
    floatRates = GetFloatRateExtraData(data.sheet_by_name(u'2_现金流示例'))
    riskfreeRate = GetBaseEarningRate(data.sheet_by_name(u'3_特定估值日期无风险利率'), None)
    # closingPrice =  GetClosingPrice(baseData, None)
    secutiries = {}
    for row in range(1, baseData.nrows):
        value = baseData.row_values(row)
        stay = value[29]
        if stay == '保留':
            code, name, couponType, frequency = value[1], value[2], value[6], int(value[7])
            rateType, couponRate, rateInfo = value[8], float(value[9]), value[10]
            date1 = xlrd.xldate_as_datetime(baseData.cell_value(row, 12), 0).strftime('%Y/%m/%d')    # 转换日期格式
            date2 = xlrd.xldate_as_datetime(baseData.cell_value(row, 13), 0).strftime('%Y/%m/%d')
            carrayDate, maturitydate = Timestamp(date1), Timestamp(date2)
            closingPrice = value[26]
            secutiries[code] = Secutiry(code, name, 
                couponType, couponRate / 100, 
                rateInfo, rateType, frequency, 
                carrayDate, maturitydate,
                value[30], None, closingPrice)
            # 算好利率
            # secutiries[code].GetRateArray(floatRates)
    return data, secutiries, floatRates, riskfreeRate#, closingPrice

# 现金流保存在result.txt文件中
def WriteCashflowToFile(secutiries, outputDir, date: str, floatRates):
    # 如果输入all那就把债券自发行日开始的所有现金流算出来
    # if mode == 'all':
    #     with open(outputDir, 'w', encoding='utf-8') as file:
    #         for key in secutiries:
    #             code = secutiries[key].code
    #             date = secutiries[key].carrayDate
    #             timeStamps, cashFlow = secutiries[code].CalcCashFlow(date)
    #             file.write(code + ' ' + str(date) + '\n')
    #             file.write(str(timeStamps) + '\n')
    #             file.write(str([float('{:.2f}'.format(i)) for i in cashFlow]) + '\n')
    #         file.close()
    # else:
    date = Timestamp(date)
    with open(outputDir, 'w', encoding='utf-8') as file:
        for key in secutiries:
            code = secutiries[key].code
            # 根据查询日期变更利率数组
            #if secutiries[key].rateType == u'浮动利率':
            secutiries[key].EstimateRateArray(date, floatRates)
            timeStamps, cashFlow = secutiries[code].CalcCashFlow(date)
            file.write(code + ' ' + str(date) + '\n')
            file.write(str(timeStamps) + '\n')
            file.write(str([float('{:.2f}'.format(i)) for i in cashFlow]) + '\n')
        file.close()

def GetRisk(secutiries, date, riskfreeRate, closingPrice):
    date = Timestamp(date)
    riskDict = {}
    riskDict2 = {}
    for code in secutiries:
        timeStamps, cashflow = secutiries[code].CalcCashFlow(date)
        cashflow = [float('{:.2f}'.format(i)) for i in cashflow]

        # 更新期限
        secutiries[code].UpdateMaturity(date.__str__())
        # baseRate = riskfreeRate.GetRate(secutiries[code].maturity)  # 无风险利率
        price = secutiries[code].closingPrice                       # 获得收盘全价

        risk = CalcZSpread(price, date.__str__(), timeStamps, cashflow, riskfreeRate)

        risk = risk * 10000
        riskDict2[code] = risk
        if secutiries[code].category in riskDict:
            riskDict[secutiries[code].category].append(risk)
        else:
            riskDict[secutiries[code].category] = [risk]
    return riskDict, riskDict2

# 输入债券代码和估值日期返回（时间戳，现金流）
def GetCashflow(secutiries, code, date):
    if code in secutiries:
        # timeStamps, cashFlow = secutiries[code].CalcCashFlow(Timestamp(date))
        return secutiries[code].CalcCashFlow(Timestamp(date))
    else:
        print('未查询到输入的债券！')

def ReadRisk(filepath):
    print('读取risk文件...')
    risk_code = {}
    with open(filepath, 'r', encoding='utf-8') as file:
        line = file.readline()
        while line:
            info = line.split(' ')
            risk_code[info[0]] = float(info[1])
            line = file.readline()
    file.close()
    return risk_code

def WriteRisk(filepath, risk_code):
    with open(filepath, 'w', encoding='utf-8') as file:
        for code in risk_code:
            if risk_code[code] < 0:
                continue
            file.write(code + ' ' + str(risk_code(code)) + '\n')
    file.close()

def WriteExcel(filepath, data, risk_code):
    # 使用openpyxl写入
    workbook = openpyxl.load_workbook(filepath)
    worksheet = workbook.worksheets[2]
    # old = copy(data)
    baseData = data.sheet_by_name(u'1_特定估值日期数据')
    HYInfo = []
    IGInfo = []

    for row in range(1, baseData.nrows):
        value = baseData.row_values(row)
        stay = value[29]
        if stay == '保留':
            if risk_code[value[1]] < 0:
                worksheet.cell(row + 1, 30, '剔除')
                continue
            worksheet.cell(row + 1, 32, risk_code[value[1]])
            # 把原值不动地填进去，之后再改
            worksheet.cell(row + 1, 25, value[23])
            # if value[23] == 'HY':
            #     if risk_code[value[1]] > 99:
            #         worksheet.cell(row + 1, 25, '其他级')
            #     else:
            #         worksheet.cell(row + 1, 25, 'IG' )
            # elif value[23] == 'IG':
            #     if risk_code[value[1]] > 99:
            #         worksheet.cell(row + 1, 25, 'HY')
            #     else:
            #         worksheet.cell(row + 1, 25, 'IG')
            if value[23] == 'HY':
                HYInfo.append([row + 1, 25, risk_code[value[1]]])
            elif value[23] == 'IG':
                IGInfo.append([row + 1, 25, risk_code[value[1]]])
    HYInfo.sort(key=TakeThird)
    IGInfo.sort(key=TakeThird)
    for i in range(0, int(0.01 * len(HYInfo) + 0.5)):
        row, col, value = HYInfo[i][0], HYInfo[i][1], 'IG'
        worksheet.cell(row, col, value)
    for i in range(int(0.99 * len(HYInfo) + 0.5), len(HYInfo)):
        row, col, value = HYInfo[i][0], HYInfo[i][1], '其他级'
        worksheet.cell(row, col, value)
    for i in range(int(0.99 * len(IGInfo) + 0.5), len(IGInfo)):
        row, col, value = IGInfo[i][0], IGInfo[i][1], 'HY'
        worksheet.cell(row, col, value)
    
    workbook.save(filename=filepath)

if __name__ == '__main__':
    data, secutiries, floatRates, riskfreeRates = ReadExcelData('C:\\Users\\86183\\Documents\\temp\\信用利差计算-20110131_old.xlsx')
    WriteCashflowToFile(secutiries, 'result_20110131.txt', '2011/1/31', floatRates)

    # risk_catagory: 把债券按风险因子分类，risk_code：按code查询债券利差
    risk_catagory, risk_code = GetRisk(secutiries, '2011/1/31', riskfreeRates, None)
    # risk_code = ReadRisk('risk.txt')
    WriteRisk('risk.txt', risk_code)
    WriteExcel('C:\\Users\\86183\\Documents\\temp\\信用利差计算-20110131.xlsx', data, risk_code)

    for key in risk_catagory:
        print(key, np.median(risk_catagory[key]))

    # while True:
    #     code = input('请输入债券代码：')
    #     date = input('请输入估值日期：')
    #     timeStamps, cashflow = GetCashflow(secutiries, code, date)
    #     cashflow = [float('{:.2f}'.format(i)) for i in cashflow]    # 保留两位小数
    #     print(timeStamps)
    #     print(cashflow)
